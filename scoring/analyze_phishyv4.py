import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone, timedelta
import os, argparse

# ── Arguments ──────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser(description="Analyze PhishyMailbox CSV export")
parser.add_argument("csv")
parser.add_argument("--after",      default=None, help="Only include participants who started after this Eastern time. Format: YYYY-MM-DD or 'YYYY-MM-DD HH:MM'")
parser.add_argument("--before",     default=None, help="Only include participants who started before this Eastern time.")
parser.add_argument("--conditions", default=None, help="Path to a CSV with columns RandID and Condition")
parser.add_argument("--output",     default="phishy_analysis.xlsx")
args = parser.parse_args()

EASTERN = timezone(timedelta(hours=-4))  # EDT; change to -5 for EST in winter

def to_eastern(dt):
    return dt.astimezone(EASTERN) if pd.notna(dt) else pd.NaT

def parse_eastern(s):
    s = s.strip()
    try:    dt = datetime.strptime(s, "%Y-%m-%d %H:%M")
    except: dt = datetime.strptime(s, "%Y-%m-%d")
    return dt.replace(tzinfo=EASTERN)

def secs_to_str(s):
    if s is None or pd.isna(s) or s < 0: return ""
    s = int(s)
    return f"{s//60}m {s%60}s"

def avg_secs_to_str(vals):
    vals = [v for v in vals if v is not None and not pd.isna(v) and v >= 0]
    if not vals: return ""
    return secs_to_str(sum(vals) / len(vals))

def make_link_label(url, text):
    url_str  = str(url).strip() if pd.notna(url) else ""
    text_str = str(text).strip().replace("\n", " ").replace("\r", "") if pd.notna(text) else ""
    text_part = text_str[:15].strip()
    url_part  = url_str[:15].strip()
    return f"{text_part} | {url_part}" if text_part else url_part

# ── Load data ──────────────────────────────────────────────────────────────────
df = pd.read_csv(args.csv)
df["At"] = pd.to_datetime(df["At"], utc=True, format="mixed")

# ── Conditions ─────────────────────────────────────────────────────────────────
condition_map = {}
all_conditions = []
if args.conditions:
    cdf = pd.read_csv(args.conditions)
    cdf.columns = [c.strip() for c in cdf.columns]
    for _, row in cdf.iterrows():
        condition_map[str(row["RandID"]).strip()] = str(row["Condition"]).strip()
    all_conditions = sorted(set(condition_map.values()))
    print(f"  Conditions loaded for {len(condition_map)} participants: {all_conditions}")

# ── Date filter ────────────────────────────────────────────────────────────────
filter_after  = parse_eastern(args.after)  if args.after  else None
filter_before = parse_eastern(args.before) if args.before else None
if filter_after or filter_before:
    started = df[df["Type"] == "started"][["Code", "At"]]
    valid = set()
    for _, r in started.iterrows():
        t = r["At"]
        if filter_after  and t < filter_after:  continue
        if filter_before and t > filter_before: continue
        valid.add(r["Code"])
    df = df[df["Code"].isin(valid)]
    print(f"  Date filter applied — {len(valid)} participants match.")

# ── Drop incomplete participants ───────────────────────────────────────────────
finished = df[df["Type"] == "finished"]
valid_codes = set(finished["Code"].unique())
df = df[df["Code"].isin(valid_codes)]
print(f"  Removed incomplete participants — {len(valid_codes)} participants completed.")

def is_phish(e): return str(e).startswith("P")
def is_legit(e): return str(e).startswith("L")
def is_spam(e):  return str(e).startswith("S")

# ── Email discovery ────────────────────────────────────────────────────────────
email_order, seen = [], set()
for _, row in df[df["Email-Backoffice-ID"].notna() & (df["Email-Backoffice-ID"] != "")].iterrows():
    eid = row["Email-Backoffice-ID"]
    if eid not in seen:
        email_order.append(eid)
        seen.add(eid)

participants = sorted(df["Code"].unique(), key=str)

# ── Discover unique hover links per email ──────────────────────────────────────
hovers_df = df[df["Type"] == "email-link-hover"]
email_hover_links = {}  # email_id -> [(label, url), ...]
for email_id in email_order:
    eh = hovers_df[hovers_df["Email-Backoffice-ID"] == email_id].sort_values("At")
    seen_urls, links = set(), []
    for _, row in eh.iterrows():
        url = str(row["URL"]).strip() if pd.notna(row["URL"]) else ""
        if url and url not in seen_urls:
            seen_urls.add(url)
            links.append((make_link_label(row["URL"], row["Link Text"]), url))
    email_hover_links[email_id] = links

# ── Sub-column definitions ─────────────────────────────────────────────────────
PHISH_SUB = ["Clicked Link", "Score", "Replied", "Folder", "Time Before First Click", "Time Before Move"]
SPAM_SUB  = ["Clicked Link", "Replied", "Folder", "Time Before First Click", "Time Before Move"]
LEGIT_SUB = ["Clicked Link", "Replied", "Folder", "Time Before First Click", "Time Before Move"]

SUB_WIDTHS = {
    "Folder": 14, "Clicked Link": 12, "Replied": 10, "Score": 11,
    "Time Before First Click": 18, "Time Before Move": 14,
}
HOVER_WIDTH = 20

def get_sub_cols(email_id):
    return PHISH_SUB if is_phish(email_id) else (SPAM_SUB if is_spam(email_id) else LEGIT_SUB)

def get_all_cols(email_id):
    base  = get_sub_cols(email_id)
    hover = [f"Hover: {label}" for label, _ in email_hover_links.get(email_id, [])]
    # Insert hover cols between Folder and Time Before First Click
    folder_idx = base.index("Folder") + 1
    return base[:folder_idx] + hover + base[folder_idx:]

# ── Build column index map ─────────────────────────────────────────────────────
meta_cols = 8
email_col_map = {}
ci = meta_cols + 1
for email_id in email_order:
    email_col_map[email_id] = {}
    for s in get_all_cols(email_id):
        email_col_map[email_id][s] = ci
        ci += 1

# ── Build participant data rows ────────────────────────────────────────────────
rows = []
for code in participants:
    p = df[df["Code"] == code].sort_values("At")
    rid = str(p["RandID"].dropna().iloc[0]).strip() if p["RandID"].notna().any() else ""
    condition = condition_map.get(rid, "") if condition_map else ""

    started_rows  = p[p["Type"] == "started"]
    finished_rows = p[p["Type"] == "finished"]
    started_at  = started_rows["At"].iloc[0]  if not started_rows.empty  else pd.NaT
    finished_at = finished_rows["At"].iloc[0] if not finished_rows.empty else pd.NaT

    started_et_str = to_eastern(started_at).strftime("%Y-%m-%d %H:%M ET") if pd.notna(started_at) else ""
    total_mins = round((finished_at - started_at).total_seconds() / 60, 1) \
        if pd.notna(started_at) and pd.notna(finished_at) else ""

    row = {
        "Participant Code":       code,
        "RandID":                 rid,
        "Condition":              condition,
        "Started At (ET)":        started_et_str,
        "Finished":               "Yes" if pd.notna(finished_at) else "No",
        "Total Time (mins)":      total_mins,
        "Phish Score":            "",
        "Legit Flagged as Phish": "",
    }

    phish_correct = phish_total = legit_flagged = 0

    for email_id in email_order:
        e = p[p["Email-Backoffice-ID"] == email_id].sort_values("At")
        sub_cols = get_sub_cols(email_id)

        if e.empty:
            for s in get_all_cols(email_id):
                row[f"{email_id} | {s}"] = "not assigned" if s == "Folder" else ""
            continue

        moved  = e[e["Type"] == "email-moved"]
        views  = e[e["Type"] == "email-view"]
        clicks = e[e["Type"] == "email-link-click"]
        replies= e[e["Type"] == "email-send-reply"]

        folder       = moved["To Folder"].iloc[-1] if not moved.empty else "not moved"
        first_view   = views["At"].iloc[0]          if not views.empty  else pd.NaT
        first_click  = clicks["At"].iloc[0]         if not clicks.empty else pd.NaT
        clicked_link = "Yes" if not clicks.empty else "No"
        replied      = "Yes" if not replies.empty else "No"

        click_secs = (first_click - first_view).total_seconds() \
            if pd.notna(first_view) and pd.notna(first_click) else None
        move_secs  = (moved["At"].iloc[-1] - first_view).total_seconds() \
            if pd.notna(first_view) and not moved.empty else None

        row[f"{email_id} | Folder"]                  = folder
        row[f"{email_id} | Clicked Link"]            = clicked_link
        row[f"{email_id} | Time Before First Click"] = secs_to_str(click_secs)
        row[f"{email_id} | Time Before Move"]        = secs_to_str(move_secs)
        row[f"{email_id} | _click_secs"]             = click_secs
        row[f"{email_id} | _move_secs"]              = move_secs
        row[f"{email_id} | Replied"]                 = replied

        if is_phish(email_id):
            phish_total += 1
            score = "Correct" if clicked_link == "No" else "Incorrect"
            if score == "Correct": phish_correct += 1
            row[f"{email_id} | Score"] = score
        elif is_legit(email_id) and folder == "Phishing":
            legit_flagged += 1

        # Hover times per link
        for label, url in email_hover_links.get(email_id, []):
            link_hovers = e[(e["Type"] == "email-link-hover") & (e["URL"] == url)]
            if not link_hovers.empty and pd.notna(first_view):
                hover_secs = (link_hovers["At"].iloc[0] - first_view).total_seconds()
                row[f"{email_id} | Hover: {label}"] = secs_to_str(hover_secs)
            else:
                row[f"{email_id} | Hover: {label}"] = ""

    row["Phish Score"]            = f"{phish_correct}/{phish_total}" if phish_total else ""
    row["Legit Flagged as Phish"] = legit_flagged
    rows.append(row)

result_df = pd.DataFrame(rows)

# Reorder columns — exclude internal _secs columns
meta_col_names = ["Participant Code", "RandID", "Condition", "Started At (ET)", "Finished",
                  "Total Time (mins)", "Phish Score", "Legit Flagged as Phish"]
email_cols = []
for email_id in email_order:
    for s in get_all_cols(email_id):
        email_cols.append(f"{email_id} | {s}")
result_df = result_df[meta_col_names + email_cols]

n_participants = len(rows)
data_start_row = 3
data_end_row   = 2 + n_participants
totals_start   = data_end_row + 1

all_folders = sorted(set(df[df["Type"] == "email-moved"]["To Folder"].dropna().unique()))

# ── Pre-compute email data lookup ──────────────────────────────────────────────
email_data_lookup = {}
for email_id in email_order:
    p_email = df[df["Email-Backoffice-ID"] == email_id]
    lookup = {}
    for code in participants:
        p_code  = p_email[p_email["Code"] == code]
        p_full  = df[df["Code"] == code]
        rid     = str(p_full["RandID"].dropna().iloc[0]).strip() if p_full["RandID"].notna().any() else ""
        cond    = condition_map.get(rid, "") if condition_map else ""
        clicked = not p_code[p_code["Type"] == "email-link-click"].empty
        replied = not p_code[p_code["Type"] == "email-send-reply"].empty
        moved   = p_code[p_code["Type"] == "email-moved"]
        views   = p_code[p_code["Type"] == "email-view"]
        clicks  = p_code[p_code["Type"] == "email-link-click"]
        folder  = moved["To Folder"].iloc[-1] if not moved.empty else None
        fv      = views["At"].iloc[0]  if not views.empty  else pd.NaT
        fc      = clicks["At"].iloc[0] if not clicks.empty else pd.NaT
        click_s = (fc - fv).total_seconds() if pd.notna(fv) and pd.notna(fc) else None
        move_s  = (moved["At"].iloc[-1] - fv).total_seconds() if pd.notna(fv) and not moved.empty else None

        # Hover times per link
        hover_times = {}
        for label, url in email_hover_links.get(email_id, []):
            lh = p_code[(p_code["Type"] == "email-link-hover") & (p_code["URL"] == url)]
            if not lh.empty and pd.notna(fv):
                hover_times[label] = (lh["At"].iloc[0] - fv).total_seconds()
            else:
                hover_times[label] = None

        lookup[code] = {
            "clicked": clicked, "replied": replied, "folder": folder,
            "condition": cond, "click_secs": click_s, "move_secs": move_s,
            "hover_times": hover_times,
        }
    email_data_lookup[email_id] = lookup

# Per-participant phish scores and condition mapping
participant_phish_scores = {}
p_rid_map = {}
for row in rows:
    code = row["Participant Code"]
    s = row["Phish Score"]
    if s and "/" in str(s):
        num, den = str(s).split("/")
        participant_phish_scores[code] = int(num) / int(den) if int(den) > 0 else None
    else:
        participant_phish_scores[code] = None
for code in participants:
    p_full = df[df["Code"] == code]
    rid = str(p_full["RandID"].dropna().iloc[0]).strip() if p_full["RandID"].notna().any() else ""
    p_rid_map[code] = condition_map.get(rid, "") if condition_map else ""

# ── Styles ─────────────────────────────────────────────────────────────────────
HEADER_FILL    = PatternFill("solid", fgColor="1F4E79")
PHISH_FILL     = PatternFill("solid", fgColor="C00000")
SPAM_FILL      = PatternFill("solid", fgColor="833C00")
LEGIT_FILL     = PatternFill("solid", fgColor="375623")
SUMMARY_FILL   = PatternFill("solid", fgColor="404040")
TOTALS_HDR     = PatternFill("solid", fgColor="2E4057")
HOVER_FILL     = PatternFill("solid", fgColor="31628A")
ALT_FILL       = PatternFill("solid", fgColor="EBF3FB")
WHITE_FILL     = PatternFill("solid", fgColor="FFFFFF")
CORRECT_FILL   = PatternFill("solid", fgColor="E2EFDA")
INCORRECT_FILL = PatternFill("solid", fgColor="FFDDD9")
LEGIT_FLAG     = PatternFill("solid", fgColor="FFF2CC")
SPAM_FLAG      = PatternFill("solid", fgColor="FCE4D6")
TOTALS_ALT     = PatternFill("solid", fgColor="E8EDF2")
COND_FILLS     = [
    PatternFill("solid", fgColor="1F4E79"), PatternFill("solid", fgColor="375623"),
    PatternFill("solid", fgColor="833C00"), PatternFill("solid", fgColor="2E4057"),
    PatternFill("solid", fgColor="C00000"), PatternFill("solid", fgColor="404040"),
    PatternFill("solid", fgColor="7F6000"), PatternFill("solid", fgColor="4B0082"),
]

WHITE_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT   = Font(name="Arial", size=10)
BOLD_FONT   = Font(name="Arial", bold=True, size=10)
GREEN_FONT  = Font(name="Arial", bold=True, color="375623", size=10)
RED_FONT    = Font(name="Arial", bold=True, color="C00000", size=10)
AMBER_FONT  = Font(name="Arial", bold=True, color="7F6000", size=10)
ORANGE_FONT = Font(name="Arial", bold=True, color="833C00", size=10)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
BLACK_SIDE  = Side(style="medium", color="000000")

cols = list(result_df.columns)

os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Participant Summary"

# ── Row 1: email group headers ─────────────────────────────────────────────────
ws.append([""] * len(cols))
col_idx = meta_cols + 1
for email_id in email_order:
    all_sub = get_all_cols(email_id)
    end_col = col_idx + len(all_sub) - 1
    ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=end_col)
    cell = ws.cell(row=1, column=col_idx)
    cell.value = email_id
    cell.fill  = PHISH_FILL if is_phish(email_id) else (SPAM_FILL if is_spam(email_id) else LEGIT_FILL)
    cell.font  = WHITE_FONT
    cell.alignment = CENTER
    col_idx += len(all_sub)

# ── Row 2: sub-headers ─────────────────────────────────────────────────────────
sub_headers = cols[:meta_cols] + [c.split(" | ")[1] for c in cols[meta_cols:]]
ws.append(sub_headers)
for i in range(1, len(cols) + 1):
    cell = ws.cell(row=2, column=i)
    sub  = cols[i - 1].split(" | ")[1] if " | " in cols[i - 1] else ""
    if i <= meta_cols:
        cell.fill = SUMMARY_FILL
    elif sub.startswith("Hover:"):
        cell.fill = HOVER_FILL
    else:
        cell.fill = HEADER_FILL
    cell.font = WHITE_FONT
    cell.alignment = CENTER

# ── Data rows ──────────────────────────────────────────────────────────────────
for r_idx, row_data in enumerate(result_df.itertuples(index=False), start=3):
    base_fill = ALT_FILL if r_idx % 2 == 0 else WHITE_FILL
    for c_idx, val in enumerate(row_data, start=1):
        cell = ws.cell(row=r_idx, column=c_idx)
        cell.value = val
        cell.alignment = CENTER if c_idx > meta_cols else LEFT
        col_name   = cols[c_idx - 1]
        email_name = col_name.split(" | ")[0] if " | " in col_name else ""
        sub        = col_name.split(" | ")[1] if " | " in col_name else ""

        if "| Score" in col_name:
            if val == "Correct":     cell.fill, cell.font = CORRECT_FILL, GREEN_FONT
            elif val == "Incorrect": cell.fill, cell.font = INCORRECT_FILL, RED_FONT
            else:                    cell.fill, cell.font = base_fill, BODY_FONT
        elif "| Clicked Link" in col_name and val == "Yes":
            if is_phish(email_name):  cell.fill, cell.font = INCORRECT_FILL, RED_FONT
            elif is_spam(email_name): cell.fill, cell.font = SPAM_FLAG, ORANGE_FONT
            else:                     cell.fill, cell.font = LEGIT_FLAG, AMBER_FONT
        elif "| Replied" in col_name and val == "Yes":
            cell.fill, cell.font = LEGIT_FLAG, AMBER_FONT
        elif col_name == "Legit Flagged as Phish" and isinstance(val, int) and val > 0:
            cell.fill, cell.font = INCORRECT_FILL, RED_FONT
        elif sub.startswith("Hover:") and val:
            cell.fill = PatternFill("solid", fgColor="D6E8F5")
            cell.font = BODY_FONT
        elif c_idx <= meta_cols:
            cell.fill, cell.font = base_fill, BOLD_FONT
        else:
            cell.fill, cell.font = base_fill, BODY_FONT

# ── Totals section ─────────────────────────────────────────────────────────────
totals_metrics = [
    ("# Clicked",             "clicked"),
    ("# Not Clicked",         "not_clicked"),
    ("% Clicked",             "pct_clicked"),
    ("",                      "blank1"),
]
for folder in all_folders:
    totals_metrics.append((f"# in {folder}", f"folder_{folder}"))
totals_metrics += [
    ("",                      "blank2"),
    ("# Replied",             "replied"),
    ("# Not Replied",         "not_replied"),
    ("% Replied",             "pct_replied"),
    ("",                      "blank3"),
    ("Avg Time Before Click", "avg_click_time"),
    ("Avg Time Before Move",  "avg_move_time"),
]

ws.cell(row=totals_start, column=1, value="TOTALS").fill = TOTALS_HDR
ws.cell(row=totals_start, column=1).font = WHITE_FONT
ws.cell(row=totals_start, column=1).alignment = LEFT

metric_row = {}
r = totals_start + 1
for label, key in totals_metrics:
    metric_row[key] = r
    if label:
        c = ws.cell(row=r, column=1)
        c.value, c.font, c.fill, c.alignment = label, BOLD_FONT, TOTALS_ALT if r % 2 == 0 else WHITE_FILL, LEFT
    r += 1

def write_cell(ws, row, col, val, fmt=None):
    c = ws.cell(row=row, column=col)
    c.value, c.font, c.alignment = val, BODY_FONT, CENTER
    if fmt: c.number_format = fmt

for email_id in email_order:
    sub_cols    = get_sub_cols(email_id)
    has_replied = "Replied" in sub_cols
    lookup      = email_data_lookup[email_id]
    ci_click    = email_col_map[email_id]["Clicked Link"]
    ci_folder   = email_col_map[email_id]["Folder"]
    ci_reply    = email_col_map[email_id]["Replied"] if has_replied else None
    ci_tclick   = email_col_map[email_id]["Time Before First Click"]
    ci_tmove    = email_col_map[email_id]["Time Before Move"]

    n_clicked = sum(1 for v in lookup.values() if v["clicked"])
    n_total   = len(lookup)
    pct_click = round(n_clicked / n_total, 4) if n_total > 0 else ""

    write_cell(ws, metric_row["clicked"],     ci_click, n_clicked)
    write_cell(ws, metric_row["not_clicked"], ci_click, n_total - n_clicked)
    write_cell(ws, metric_row["pct_clicked"], ci_click, pct_click if pct_click != "" else "", "0%" if pct_click != "" else None)

    for folder in all_folders:
        write_cell(ws, metric_row[f"folder_{folder}"], ci_folder,
                   sum(1 for v in lookup.values() if v["folder"] == folder))

    if has_replied and ci_reply:
        n_rep = sum(1 for v in lookup.values() if v["replied"])
        pct_r = round(n_rep / n_total, 4) if n_total > 0 else ""
        write_cell(ws, metric_row["replied"],     ci_reply, n_rep)
        write_cell(ws, metric_row["not_replied"], ci_reply, n_total - n_rep)
        write_cell(ws, metric_row["pct_replied"], ci_reply, pct_r if pct_r != "" else "", "0%" if pct_r != "" else None)

    write_cell(ws, metric_row["avg_click_time"], ci_tclick,
               avg_secs_to_str([v["click_secs"] for v in lookup.values()]))
    write_cell(ws, metric_row["avg_move_time"], ci_tmove,
               avg_secs_to_str([v["move_secs"] for v in lookup.values()]))

# ── Column widths + borders ────────────────────────────────────────────────────
for i, w in enumerate([18, 10, 14, 22, 10, 16, 14, 20], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
for i, col in enumerate(cols[meta_cols:], start=meta_cols + 1):
    sub = col.split(" | ")[1]
    ws.column_dimensions[get_column_letter(i)].width = HOVER_WIDTH if sub.startswith("Hover:") else SUB_WIDTHS.get(sub, 14)

ws.row_dimensions[1].height = 60
ws.row_dimensions[2].height = 40
ws.freeze_panes = f"{get_column_letter(meta_cols + 1)}3"

for email_id in email_order:
    all_sub   = get_all_cols(email_id)
    first_col = email_col_map[email_id][all_sub[0]]
    last_col  = email_col_map[email_id][all_sub[-1]]
    for row in range(1, ws.max_row + 1):
        lc = ws.cell(row=row, column=first_col)
        rc = ws.cell(row=row, column=last_col)
        lc.border = Border(left=BLACK_SIDE, right=lc.border.right if lc.border else Side())
        rc.border = Border(left=rc.border.left if rc.border else Side(), right=BLACK_SIDE)

# ── Helper: write summary value into a cell ────────────────────────────────────
def write_summary_value(cell, key, lookup, sub, sub_cols):
    if sub and sub not in sub_cols and not sub.startswith("Hover:"):
        cell.value = "N/A"
        return
    n_total = len(lookup)
    if key == "clicked":
        cell.value = sum(1 for v in lookup.values() if v["clicked"])
    elif key == "not_clicked":
        cell.value = n_total - sum(1 for v in lookup.values() if v["clicked"])
    elif key == "pct_clicked":
        n = sum(1 for v in lookup.values() if v["clicked"])
        cell.value = round(n / n_total, 4) if n_total > 0 else ""
        if cell.value != "": cell.number_format = "0%"
    elif key.startswith("folder_"):
        folder = key[len("folder_"):]
        count  = sum(1 for v in lookup.values() if v["folder"] == folder)
        pct    = round(count / n_total * 100) if n_total > 0 else 0
        cell.value = f"{count} ({pct}%)"
    elif key == "replied":
        cell.value = sum(1 for v in lookup.values() if v["replied"])
    elif key == "not_replied":
        cell.value = n_total - sum(1 for v in lookup.values() if v["replied"])
    elif key == "pct_replied":
        n = sum(1 for v in lookup.values() if v["replied"])
        cell.value = round(n / n_total, 4) if n_total > 0 else ""
        if cell.value != "": cell.number_format = "0%"
    elif key == "avg_click_time":
        cell.value = avg_secs_to_str([v["click_secs"] for v in lookup.values()])
    elif key == "avg_move_time":
        cell.value = avg_secs_to_str([v["move_secs"] for v in lookup.values()])

# ── Build summary metrics list ─────────────────────────────────────────────────
def build_summary_metrics(all_folders):
    m = [
        ("Clicked", "# Clicked",             "clicked",        "Clicked Link"),
        ("Clicked", "# Not Clicked",          "not_clicked",    "Clicked Link"),
        ("Clicked", "% Clicked",              "pct_clicked",    "Clicked Link"),
    ]
    for folder in all_folders:
        m.append(("Folders", f"# in {folder}", f"folder_{folder}", "Folder"))
    m += [
        ("Replies", "# Replied",              "replied",        "Replied"),
        ("Replies", "# Not Replied",           "not_replied",    "Replied"),
        ("Replies", "% Replied",               "pct_replied",    "Replied"),
        ("Time",    "Avg Time Before Click",   "avg_click_time", "Time Before First Click"),
        ("Time",    "Avg Time Before Move",    "avg_move_time",  "Time Before Move"),
    ]
    return m

summary_metrics = build_summary_metrics(all_folders)

def style_summary_cat(cell, cat):
    cell.font = WHITE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.fill = TOTALS_HDR if cat == "Clicked" else \
                (LEGIT_FILL if cat == "Replies" else \
                (PHISH_FILL if cat == "Score" else SUMMARY_FILL))

def add_section_borders(sheet, section_last, max_col):
    for cat, last_row in section_last.items():
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=last_row, column=col)
            cell.border = Border(
                left=cell.border.left, right=cell.border.right, top=cell.border.top,
                bottom=Side(style="medium", color="000000")
            )

# ── Email Summary tab ──────────────────────────────────────────────────────────
ss = wb.create_sheet("Email Summary")
ss.cell(row=1, column=1, value="Category").fill = SUMMARY_FILL
ss.cell(row=1, column=1).font = WHITE_FONT
ss.cell(row=1, column=1).alignment = CENTER
ss.cell(row=1, column=2, value="Metric").fill = SUMMARY_FILL
ss.cell(row=1, column=2).font = WHITE_FONT
ss.cell(row=1, column=2).alignment = CENTER

for e_idx, email_id in enumerate(email_order, start=3):
    c = ss.cell(row=1, column=e_idx)
    c.value = email_id
    c.fill  = PHISH_FILL if is_phish(email_id) else (SPAM_FILL if is_spam(email_id) else LEGIT_FILL)
    c.font  = WHITE_FONT
    c.alignment = CENTER

cat_rows, sect_last = {}, {}
for r_idx, (cat, label, key, sub) in enumerate(summary_metrics, start=2):
    if cat not in cat_rows: cat_rows[cat] = [r_idx, r_idx]
    else: cat_rows[cat][1] = r_idx
    sect_last[cat] = r_idx

for r_idx, (cat, label, key, sub) in enumerate(summary_metrics, start=2):
    base = TOTALS_ALT if r_idx % 2 == 0 else WHITE_FILL
    ss.cell(row=r_idx, column=1).fill = base
    ss.cell(row=r_idx, column=1).alignment = CENTER
    cb = ss.cell(row=r_idx, column=2)
    cb.value, cb.font, cb.fill, cb.alignment = label, BOLD_FONT, base, LEFT

    for e_idx, email_id in enumerate(email_order, start=3):
        sub_cols = get_sub_cols(email_id)
        lookup   = email_data_lookup[email_id]
        c = ss.cell(row=r_idx, column=e_idx)
        c.fill, c.alignment, c.font = base, CENTER, BODY_FONT
        write_summary_value(c, key, lookup, sub, sub_cols)

for cat, (r_start, r_end) in cat_rows.items():
    if r_start < r_end:
        ss.merge_cells(start_row=r_start, start_column=1, end_row=r_end, end_column=1)
    style_summary_cat(ss.cell(row=r_start, column=1), cat)
    ss.cell(row=r_start, column=1).value = cat

add_section_borders(ss, sect_last, len(email_order) + 2)
ss.column_dimensions["A"].width = 12
ss.column_dimensions["B"].width = 22
for i in range(3, len(email_order) + 3):
    ss.column_dimensions[get_column_letter(i)].width = 22
ss.row_dimensions[1].height = 60
ss.freeze_panes = "C2"

# ── By Condition tab ───────────────────────────────────────────────────────────
if args.conditions and all_conditions:
    cs = wb.create_sheet("By Condition")
    n_conds = len(all_conditions)

    cs.cell(row=1, column=1, value="Category").fill = SUMMARY_FILL
    cs.cell(row=1, column=1).font = WHITE_FONT
    cs.cell(row=1, column=1).alignment = CENTER
    cs.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    cs.cell(row=1, column=2, value="Metric").fill = SUMMARY_FILL
    cs.cell(row=1, column=2).font = WHITE_FONT
    cs.cell(row=1, column=2).alignment = CENTER
    cs.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

    col_start = 3
    email_cond_col_map = {}
    for email_id in email_order:
        end_col = col_start + n_conds - 1
        cs.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=end_col)
        cell = cs.cell(row=1, column=col_start)
        cell.value = email_id
        cell.fill  = PHISH_FILL if is_phish(email_id) else (SPAM_FILL if is_spam(email_id) else LEGIT_FILL)
        cell.font  = WHITE_FONT
        cell.alignment = CENTER
        email_cond_col_map[email_id] = {}
        for c_idx, cond in enumerate(all_conditions):
            col  = col_start + c_idx
            cell = cs.cell(row=2, column=col)
            cell.value = f"Cond {cond}"
            cell.fill  = COND_FILLS[c_idx % len(COND_FILLS)]
            cell.font  = WHITE_FONT
            cell.alignment = CENTER
            email_cond_col_map[email_id][cond] = col
        col_start += n_conds

    total_data_cols = col_start - 1

    cat_rows_c, sect_last_c = {}, {}
    for r_idx, (cat, label, key, sub) in enumerate(summary_metrics, start=3):
        if cat not in cat_rows_c: cat_rows_c[cat] = [r_idx, r_idx]
        else: cat_rows_c[cat][1] = r_idx
        sect_last_c[cat] = r_idx

    for r_idx, (cat, label, key, sub) in enumerate(summary_metrics, start=3):
        base = TOTALS_ALT if r_idx % 2 == 0 else WHITE_FILL
        cs.cell(row=r_idx, column=1).fill = base
        cs.cell(row=r_idx, column=1).alignment = CENTER
        cb = cs.cell(row=r_idx, column=2)
        cb.value, cb.font, cb.fill, cb.alignment = label, BOLD_FONT, base, LEFT

        for email_id in email_order:
            sub_cols = get_sub_cols(email_id)
            for cond in all_conditions:
                col    = email_cond_col_map[email_id][cond]
                lookup = {c: v for c, v in email_data_lookup[email_id].items() if v["condition"] == cond}
                c = cs.cell(row=r_idx, column=col)
                c.fill, c.alignment, c.font = base, CENTER, BODY_FONT
                write_summary_value(c, key, lookup, sub, sub_cols)

    for cat, (r_start, r_end) in cat_rows_c.items():
        if r_start < r_end:
            cs.merge_cells(start_row=r_start, start_column=1, end_row=r_end, end_column=1)
        style_summary_cat(cs.cell(row=r_start, column=1), cat)
        cs.cell(row=r_start, column=1).value = cat

    add_section_borders(cs, sect_last_c, total_data_cols)

    for email_id in email_order:
        first_col = email_cond_col_map[email_id][all_conditions[0]]
        last_col  = email_cond_col_map[email_id][all_conditions[-1]]
        for row in range(1, cs.max_row + 2):
            lc = cs.cell(row=row, column=first_col)
            rc = cs.cell(row=row, column=last_col)
            lc.border = Border(left=BLACK_SIDE, right=lc.border.right if lc.border else Side())
            rc.border = Border(left=rc.border.left if rc.border else Side(), right=BLACK_SIDE)

    # ── Condition Summary table below ──────────────────────────────────────────
    sum_start = cs.max_row + 3
    hdr = cs.cell(row=sum_start, column=1, value="Condition Summary")
    hdr.fill, hdr.font, hdr.alignment = TOTALS_HDR, WHITE_FONT, LEFT
    cs.merge_cells(start_row=sum_start, start_column=1, end_row=sum_start, end_column=1 + n_conds)

    sub_hdr_row = sum_start + 1
    cs.cell(row=sub_hdr_row, column=1, value="Metric").fill = SUMMARY_FILL
    cs.cell(row=sub_hdr_row, column=1).font = WHITE_FONT
    cs.cell(row=sub_hdr_row, column=1).alignment = LEFT
    for c_idx, cond in enumerate(all_conditions, start=2):
        cell = cs.cell(row=sub_hdr_row, column=c_idx)
        cell.value, cell.fill, cell.font, cell.alignment = \
            f"Condition {cond}", COND_FILLS[(c_idx-2) % len(COND_FILLS)], WHITE_FONT, CENTER

    cond_rows = [
        ("Avg Phish Score",       "avg_phish"),
        ("Avg Time Before Click", "avg_click"),
        ("Avg Time Before Move",  "avg_move"),
    ]
    for r_off, (label, key) in enumerate(cond_rows, start=2):
        r = sum_start + r_off
        base = TOTALS_ALT if r % 2 == 0 else WHITE_FILL
        lc = cs.cell(row=r, column=1)
        lc.value, lc.font, lc.fill, lc.alignment = label, BOLD_FONT, base, LEFT
        for c_idx, cond in enumerate(all_conditions, start=2):
            cond_codes = [code for code in participants if p_rid_map.get(code) == cond]
            c = cs.cell(row=r, column=c_idx)
            c.fill, c.alignment, c.font = base, CENTER, BODY_FONT
            if key == "avg_phish":
                scores = [participant_phish_scores[code] for code in cond_codes
                          if participant_phish_scores.get(code) is not None]
                c.value = round(sum(scores)/len(scores), 4) if scores else "N/A"
                if c.value != "N/A": c.number_format = "0%"
            elif key == "avg_click":
                times = [v["click_secs"] for eid in email_order
                         for code, v in email_data_lookup[eid].items()
                         if v["condition"] == cond and v["click_secs"] is not None]
                c.value = avg_secs_to_str(times)
            elif key == "avg_move":
                times = [v["move_secs"] for eid in email_order
                         for code, v in email_data_lookup[eid].items()
                         if v["condition"] == cond and v["move_secs"] is not None]
                c.value = avg_secs_to_str(times)

    cs.column_dimensions["A"].width = 22
    cs.column_dimensions["B"].width = 14
    for i in range(3, total_data_cols + 1):
        cs.column_dimensions[get_column_letter(i)].width = 14
    cs.row_dimensions[1].height = 60
    cs.row_dimensions[2].height = 30
    cs.freeze_panes = "C3"

# ── Legend ─────────────────────────────────────────────────────────────────────
ls = wb.create_sheet("Legend")
legend_data = [
    ("Color",                     "Meaning"),
    ("Red email header",          "Phishing email (P) — scored on link clicks only"),
    ("Dark orange email header",  "Spam email (S) — not scored"),
    ("Green email header",        "Legitimate email (L) — not scored"),
    ("Green Score cell",          "Correct — did not click the link"),
    ("Red Score cell",            "Incorrect — clicked the link"),
    ("Red Clicked Link",          "Clicked a phishing link"),
    ("Light orange Clicked Link", "Clicked a spam link — flag for review"),
    ("Yellow Clicked Link",       "Clicked a legit link — informational"),
    ("Yellow Replied",            "Replied — informational"),
    ("Red Legit Flagged as Phish","Moved a legit email to the Phishing folder"),
    ("Blue Hover column",         "Time from first viewing the email to first hover on that link"),
    ("", ""),
    ("Scoring",                   "Phishing only: Correct = did NOT click. Incorrect = clicked."),
    ("Spam",                      "Not scored. Link clicks flagged for review."),
    ("Legit",                     "Not scored. Clicks and replies tracked informational."),
    ("Legit Flagged as Phish",    "Count of legit emails moved to the Phishing folder specifically."),
    ("", ""),
    ("Hover columns",             "One column per unique link per email (based on URL). Label = first 15 chars of link text + first 15 chars of URL. Value = time from first view to first hover on that link."),
    ("", ""),
    ("Totals section",            "Below participant rows. Includes avg time before click/move."),
    ("Email Summary tab",         "Overall totals per email, broken into Clicked/Folders/Replies/Time sections."),
    ("By Condition tab",          "Same metrics broken by condition. Separate table shows Avg Phish Score and Avg Time per condition."),
    ("", ""),
    ("--after 'YYYY-MM-DD'",      "Include participants who started on or after this date (Eastern)"),
    ("--before 'YYYY-MM-DD HH:MM'","Include participants who started before this date/time"),
    ("--conditions file.csv",     "CSV with RandID and Condition columns. Enables By Condition tab."),
    ("--output filename.xlsx",    "Custom output path"),
    ("Timezone",                  "EDT = UTC-4 (summer). Change EASTERN to UTC-5 for winter EST."),
]
for r, (a, b) in enumerate(legend_data, start=1):
    ls.cell(row=r, column=1, value=a).font = Font(name="Arial", bold=True, size=10)
    ls.cell(row=r, column=2, value=b).font = Font(name="Arial", size=10)
ls.column_dimensions["A"].width = 30
ls.column_dimensions["B"].width = 70

wb.save(args.output)
print(f"✓ Saved to {args.output}")
print(f"  {len(rows)} participants × {len(email_order)} emails")
print(f"  Totals section starts at row {totals_start}")
if all_conditions:
    print(f"  By Condition tab: {len(all_conditions)} conditions")
